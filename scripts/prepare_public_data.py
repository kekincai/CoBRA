"""Extract official public inputs, preserving sources and license notices. No company data."""

import hashlib
import json
import xml.etree.ElementTree as ET
from pathlib import Path
from urllib.request import urlopen
from zipfile import ZipFile

import numpy as np
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "src/cobra_web/data"
COBRA_URL = "https://www.ipa.go.jp/archive/files/000058309.zip"
BENCH_URL = (
    "https://www.ipa.go.jp/digital/software-survey/metrics/hjuojm000000c6it-att/000103288.zip"
)


def download(url, path):
    if not path.exists():
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(urlopen(url, timeout=60).read())
    return hashlib.sha256(path.read_bytes()).hexdigest()


def write(name, data):
    (OUT / name).write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n")


def main():
    path = ROOT / "data/public/cobra.zip"
    checksum = download(COBRA_URL, path)
    with ZipFile(path) as z:
        xml = z.read(next(n for n in z.namelist() if n.endswith("model_data_Open.xml")))
        license_text = z.read(
            next(n for n in z.namelist() if n.endswith("MIT-License.txt"))
        ).decode()
    (OUT / "IPA-CoBRA-LICENSE.txt").write_text(license_text)
    root = ET.fromstring(xml)
    drivers = []
    for node in root.findall("UserModel/CostFactors/Factor"):
        ratio = node.find(".//Ratio")
        drivers.append(
            {
                "id": node.attrib["id"],
                "name": node.findtext("Name"),
                "description": node.findtext("FactorDescription"),
                "levels": [node.findtext(f"LevelDescription/Level{i}") for i in range(4)],
                "experts": [
                    {
                        "name": "IPA 公開合意値",
                        "minimum": float(ratio.findtext("Best")),
                        "mode": float(ratio.findtext("MostLikely")),
                        "maximum": float(ratio.findtext("Worst")),
                    }
                ],
                "enabled": True,
            }
        )
    ps = root.find("ProjectData")
    projects = [
        {
            "id": p.attrib["id"],
            "name": p.findtext("Description"),
            "size": float(p.findtext("Size")),
            "size_unit": ps.attrib["unit_size"],
            "actual_effort": float(p.findtext("Effort")),
            "levels": {
                f.attrib["CostFactorId"]: int(f.text) for f in p.findall("FactorLevels/FactorLevel")
            },
            "source_type": "COBRA_PUBLIC_SAMPLE",
            "measurement": "IPA sample original unit",
            "project_type": "公開テストケース",
        }
        for p in ps
    ]
    write(
        "public_sample.json",
        {
            "source_url": COBRA_URL,
            "archive_sha256": checksum,
            "xml_sha256": hashlib.sha256(xml).hexdigest(),
            "source_type": "COBRA_PUBLIC_SAMPLE",
            "drivers": {"name": "IPA CoBRA 公開サンプル", "drivers": drivers},
            "projects": projects,
            "original_accuracy": {n.tag: float(n.text) for n in root.find("Accuracy")},
        },
    )
    path = ROOT / "data/benchmark/benchmark.zip"
    checksum = download(BENCH_URL, path)
    records = []
    with ZipFile(path) as z:
        for name in z.namelist():
            if not name.endswith(".xlsx"):
                continue
            import io

            book = openpyxl.load_workbook(io.BytesIO(z.read(name)), read_only=True, data_only=True)
            filename = name.encode("cp437").decode("cp932")
            industry = filename.split("_")[0]
            if industry == "本編":
                industry = "業種全体"
                terms = "\n".join(
                    str(c)
                    for row in book["はじめにお読みください"].iter_rows(values_only=True)
                    for c in row
                    if c
                )
                (OUT / "IPA-Benchmark-TERMS.txt").write_text(
                    terms
                    + "\n\n加工: 公開集計表から数値・出典セルを JSON に抽出。著作権表示: Copyright 2022 IPA。加工および本表示の責任主体: CoBRA Web contributors。加工部分について著作者人格権を行使しません。\n"
                )
            # Published summary cells, never digitized graphics or synthesized observations.
            for sheet, kind in [
                ("A1-2-1", "sloc_new"),
                ("A1-2-2", "sloc_enhancement"),
                ("A2-2-1", "fp_new"),
                ("A2-2-2", "fp_enhancement"),
                ("A3-3-8", "phase_new"),
                ("A3-3-9", "phase_enhancement"),
            ]:
                if sheet not in book.sheetnames:
                    continue
                s = book[sheet]
                for row in range(3, 8 if kind.startswith("phase") else 4):
                    records.append(
                        {
                            "industry": industry,
                            "metric": kind,
                            "label": s.cell(row, 7 if kind.startswith("phase") else 6).value,
                            "unit": "比率"
                            if kind.startswith("phase")
                            else ("FP/人時" if kind.startswith("fp") else "SLOC/人時"),
                            "n": s.cell(row, 8).value,
                            "p25": s.cell(row, 10).value,
                            "median": s.cell(row, 11).value,
                            "p75": s.cell(row, 12).value,
                            "mean": s.cell(row, 14).value,
                            "workbook": filename,
                            "sheet": sheet,
                            "cells": f"H{row}:N{row}",
                        }
                    )
            for sheet, kind in [("A3-3-1", "duration_new"), ("A3-3-2", "duration_enhancement")]:
                s = book[sheet]
                values = [
                    row[1]
                    for row in s.iter_rows(min_row=2, max_col=2, values_only=True)
                    if isinstance(row[0], (int, float)) and isinstance(row[1], (int, float))
                ]
                records.append(
                    {
                        "industry": industry,
                        "metric": kind,
                        "label": "開発期間（公開点データから集計）",
                        "unit": "月",
                        "n": len(values),
                        "p25": float(np.percentile(values, 25)),
                        "median": float(np.median(values)),
                        "p75": float(np.percentile(values, 75)),
                        "mean": float(np.mean(values)),
                        "workbook": filename,
                        "sheet": sheet,
                        "cells": f"B2:B{s.max_row}",
                        "transformation": "Non-empty published duration observations; linear quantile, no fitted-curve values.",
                    }
                )
    write(
        "benchmark.json",
        {
            "version": "IPA-2022-20230117",
            "source_type": "IPA_BENCHMARK",
            "source_url": BENCH_URL,
            "archive_sha256": checksum,
            "copyright": "Copyright 2022 IPA",
            "records": records,
        },
    )
    print(
        f"Extracted {len(drivers)} drivers, {len(projects)} public cases, {len(records)} benchmark summaries."
    )


if __name__ == "__main__":
    main()
